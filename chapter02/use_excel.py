#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
import MySQLdb


class ExcelUtils(object):
    """
    pip install pillow
    pip install openpyxl
    """

    def __init__(self):
        self.wb = Workbook()

        # 第一个活动中的sheet
        self.ws = self.wb.active
        # 创建带标题的sheet
        self.ws_two = self.wb.create_sheet("我的表单")
        # 创建不带标题的sheet
        self.ws_three = self.wb.create_sheet()

    def do_sth(self):
        # 插入数据
        self.ws['A1'] = 66
        self.ws['A2'] = '你好'
        self.ws['A3'] = datetime.now()
        # 批量插入数据
        for row in self.ws_two['A1:E6']:
            for cell in row:
                cell.value = 8
        # 对数据进行求和
        self.ws_two['G1'] = '=SUM(A1:E1)'

        # 输入图片
        img = Image('./static/temp.jpg')
        self.ws.add_image(img, 'B1')

        # 合并单元格
        self.ws.merge_cells('A4:A8')
        # 解除合并单元格
        self.ws.unmerge_cells('A4:A8')

        # 保存到磁盘
        self.wb.save("./static/text.xlsx")

    def read_xls(self):
        """
        读取excel数据
        :return: None
        """
        wb = load_workbook("./static/template.xlsx")
        names = wb.sheetnames
        # 通过下标获取cell内容
        ws = wb[names[0]]

        # 取得连接
        conn = self.get_conn()

        # 从excel中取数据
        for i, row in enumerate(ws.rows):
            if i < 2:
                continue
            year = ws['A{0}'.format(i + 1)].value
            max = ws['B{0}'.format(i + 1)].value
            avg = ws['C{0}'.format(i + 1)].value
            if year is None:
                continue

            # 写入数据到mysql
            cursor = conn.cursor()
            cursor.execute("INSERT INTO `score`(`year`,`max`,`avg`) VALUES({0}, {1}, {2})".format(year, max, avg))
            cursor.close()

            # 手动提交事务
            # conn.commit()
            # 手动回滚事务
            # conn.rollback()
            print("INSERT COMPLATE")

    def export_xls(self):
        """
        从mysql取得数据导出到excel
        :return:
        """
        # 获取数据库的连接
        conn = self.get_conn()
        cursor = conn.cursor()
        # 准备查询语句
        sql = "SELECT `year`, `max`, `avg` FROM `score`"
        # 查询数据
        cursor.execute(sql)
        rows = cursor.fetchall()
        # 循环写入到excel
        wb = Workbook()
        ws = wb.active
        for i, row in enumerate(rows):
            (
                ws['A{0}'.format(i + 1)],
                ws['B{0}'.format(i + 1)],
                ws['C{0}'.format(i + 1)]
            ) = row

        # 保存excel
        wb.save("./static/export.xlsx")

    def get_conn(self):
        """
        取得数据库连接
        :return: conn
        """
        try:
            conn = MySQLdb.connect(
                db='user_grade',
                host='localhost',
                user='pythons',
                passwd='1234',
                charset='utf8'
            )
            # 自动提交事务
            conn.autocommit(on=True)
        except BaseException as e:
            print(e)

        return conn


if __name__ == '__main__':
    client = ExcelUtils()
    # client.do_sth()
    # client.read_xls()
    client.export_xls()
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import MySQLdb
from openpyxl.chart import AreaChart, Reference


class GaokaoExport(object):

    def __init__(self):
        self.wb = load_workbook("./static/tmpl.xlsx")
        self.ws = self.wb.active
        self.ws.title = '成绩统计'
        self.ws.sheet_properties.tabColor = 'ff0000'

    def get_conn(self):
        try:
            conn = MySQLdb.connect(
                db='user_grade',
                host='localhost',
                user='pythons',
                passwd='1234',
                charset='utf8'
            )
        except Exception as e:
            print(e)
        return  conn

    def export_xls(self):
        conn = self.get_conn()
        cursor = conn.cursor()
        sql = 'SELECT `year`, `max`, `avg` FROM `score`'
        cursor.execute(sql)
        rows = cursor.fetchall()

        row_id = 10
        for i, row in enumerate(rows):
            (self.ws['C{0}'.format(row_id)],
             self.ws['D{0}'.format(row_id)],
             self.ws['E{0}'.format(row_id)]) = row
            row_id += 1

        chart = AreaChart()
        chart.title = "分数统计图"
        chart.style = 13
        chart.x_axis.title = '年份'
        chart.y_axis.title = '分数'

        x_header = Reference(self.ws, min_col=3, min_row=10, max_row=row_id-1)
        # 参数说明
        # min_col第几列开始
        # min_row第几行开始(这里取9是因为, 数据第一行默认为表头)
        # max_col第几列结束
        # max_row第几行结束
        data = Reference(self.ws, min_col=4, min_row=10, max_col=5, max_row=row_id-1)
        chart.add_data(data)
        chart.set_categories(x_header)

        self.ws.add_chart(chart, "A{0}".format(row_id+2))

        self.wb.save('./static/stats.xlsx')

if __name__ == '__main__':
    client = GaokaoExport()
    client.export_xls()